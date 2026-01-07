from django.shortcuts import get_object_or_404, render

from masters.filters import EventFilter

# Create your views here.


from .models import *
from .forms import *
from django.contrib.auth.decorators import login_required

from django.shortcuts import render, redirect
from django.urls import reverse
from django.http.response import HttpResponseRedirect

from users.permissions import *

from rest_framework.generics import ListAPIView
from django_filters.rest_framework import DjangoFilterBackend

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger


@login_required(login_url="login_admin")
def vendor_dashboard(request):

    return render(request, "vendor_dashboard.html")


@login_required(login_url="login_admin")
def register_hotel(request):

    if request.method == "POST":
        form = villa_Form()
        context = {"form": form}

        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        email = request.POST.get("email")
        mobile = request.POST.get("mobile")
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")

        if not all([first_name, last_name, email, mobile, password, confirm_password]):
            return render(
                request,
                "hotel_registration.html",
                {"error": "All fields are required."},
                context,
            )

        if password != confirm_password:
            return render(
                request,
                "hotel_registration.html",
                {"error": "Passwords do not match."},
                context,
            )

        if User.objects.filter(email=email).exists():
            return render(
                request,
                "hotel_registration.html",
                {"error": "Email already registered."},
                context,
            )

        if User.objects.filter(mobile=mobile).exists():
            return render(
                request,
                "hotel_registration.html",
                {"error": "Mobile number already registered."},
                context,
            )

        try:
            with transaction.atomic():
                # Create user
                user = User.objects.create_user(
                    email=email,
                    mobile=mobile,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    is_service_provider=True,
                    is_active=False,
                )

                form = villa_Form(request.POST, request.FILES, user=request.user)

                if form.is_valid():
                    villa_obj = form.save(commit=False)
                    if not request.user.is_superuser:
                        villa_obj.user = user
                    villa_obj.save()
                    form.save_m2m()

                    for img in request.FILES.getlist("image"):
                        VillaImage.objects.create(villa=villa_obj, image=img)

                    return redirect("list_villa")
                else:
                    # Validation failed — raise exception to rollback user
                    raise Exception(f"Form invalid: {form.errors}")

        except Exception as e:
            print("Error:", e)
            # Optional: show custom error message
            return render(
                request,
                "hotel_registration.html",
                {
                    "form": form,
                    "error": "Something went wrong during registration. Please try again.",
                },
            )

    else:
        form = villa_Form()
        from masters.models import SystemSettings

        system_settings = SystemSettings.get_settings()
        context = {
            "form": form,
            "system_settings": system_settings,
        }
        return render(request, "hotel_registration.html", context)


@login_required(login_url="login_admin")
def add_hotel(request):

    if request.method == "POST":

        form = villa_Form(request.POST, request.FILES, user=request.user)

        if form.is_valid():
            villa_obj = form.save(commit=False)
            if not request.user.is_superuser:
                villa_obj.user = request.user  # auto-assign vendor user
            villa_obj.save()
            form.save_m2m()  # Save the many-to-many relationships

            for img in request.FILES.getlist("image"):
                VillaImage.objects.create(villa=villa_obj, image=img)

            return redirect("list_villa")

        else:
            print(form.errors)
            from masters.models import SystemSettings

            system_settings = (
                SystemSettings.get_settings() if request.user.is_superuser else None
            )
            context = {
                "form": form,
                "system_settings": system_settings,
            }
            return render(request, "add_hotel.html", context)

    else:

        print(request.user)

        if request.user.is_superuser:

            print("1111")

        elif request.user.is_service_provider:

            print("----------434-----------")
        form = villa_Form()

        from masters.models import SystemSettings

        system_settings = (
            SystemSettings.get_settings() if request.user.is_superuser else None
        )

        context = {
            "form": form,
            "system_settings": system_settings,
        }

        return render(request, "add_hotel.html", context)


@login_required(login_url="login_admin")
def view_hotel(request):
    from django.db.models import Prefetch

    # Get villas for user - show first one or selected
    villa_id = request.GET.get("villa_id")
    if request.user.is_superuser:
        if villa_id:
            try:
                user_villa = (
                    villa.objects.select_related("user", "city", "property_type")
                    .prefetch_related(
                        "amenities",
                        Prefetch(
                            "rooms",
                            queryset=villa_rooms.objects.select_related(
                                "room_type"
                            ).prefetch_related("villa_amenities"),
                        ),
                    )
                    .get(id=villa_id)
                )
            except villa.DoesNotExist:
                user_villa = None
        else:
            user_villa = (
                villa.objects.select_related("user", "city", "property_type")
                .prefetch_related(
                    "amenities",
                    Prefetch(
                        "rooms",
                        queryset=villa_rooms.objects.select_related(
                            "room_type"
                        ).prefetch_related("villa_amenities"),
                    ),
                )
                .filter(is_active=True)
                .first()
            )
    else:
        # Vendor: get their villas
        villas_qs = (
            villa.objects.filter(user=request.user, is_active=True)
            .select_related("user", "city", "property_type")
            .prefetch_related(
                "amenities",
                Prefetch(
                    "rooms",
                    queryset=villa_rooms.objects.select_related(
                        "room_type"
                    ).prefetch_related("villa_amenities"),
                ),
            )
        )
        if villa_id:
            try:
                user_villa = villas_qs.get(id=villa_id)
            except villa.DoesNotExist:
                user_villa = villas_qs.first()
        else:
            user_villa = villas_qs.first()

    from masters.models import SystemSettings

    system_settings = SystemSettings.get_settings()

    context = {
        "data": user_villa,
        "system_settings": system_settings,
    }

    return render(request, "view_hotel.html", context)


@login_required(login_url="login_admin")
def update_hotel(request, villa_id):
    # Allow both admin and vendor to update villa
    # Admin can update any villa, vendor can only update their own
    if request.user.is_superuser:
        instance = villa.objects.get(id=villa_id)
    else:
        # Vendor can only update their own villa
        instance = villa.objects.get(id=villa_id, user=request.user)

    if request.method == "POST":

        instance = villa.objects.get(id=villa_id)
        forms = villa_Form(
            request.POST, request.FILES, instance=instance, user=request.user
        )

        if forms.is_valid():
            # Check if vendor is trying to change price when there are bookings
            has_bookings = False
            if not request.user.is_superuser:
                from customer.models import VillaBooking

                has_bookings = VillaBooking.objects.filter(villa=instance).exists()

                if has_bookings:
                    # Get the original price from database
                    original_price = villa.objects.get(id=villa_id).price_per_night
                    new_price = forms.cleaned_data.get("price_per_night")

                    # If price is being changed, prevent it
                    if original_price != new_price:
                        messages.error(
                            request,
                            "⚠️ You cannot change the villa price because there are existing bookings for this villa. "
                            "Please contact admin if you need to update the price.",
                        )
                        from masters.models import SystemSettings

                        system_settings = SystemSettings.get_settings()
                        context = {
                            "form": forms,
                            "existing_images": (
                                instance.images.all() if instance else None
                            ),
                            "system_settings": system_settings,
                            "has_bookings": True,
                        }
                        return render(request, "add_hotel.html", context)

            hotels = forms.save(commit=False)
            if not request.user.is_superuser:
                hotels.user = request.user  # auto-assign vendor user
                # Restore original price if vendor tried to change it
                if has_bookings:
                    hotels.price_per_night = instance.price_per_night
            hotels.save()
            forms.save_m2m()

            for img in request.FILES.getlist("image"):
                VillaImage.objects.create(villa=instance, image=img)

            if request.user.is_superuser:
                return redirect("list_villa")

            else:

                return redirect("view_villa")

        else:
            print(forms.errors)
            from masters.models import SystemSettings

            system_settings = (
                SystemSettings.get_settings() if request.user.is_superuser else None
            )
            context = {
                "form": forms,
                "existing_images": instance.images.all() if instance else None,
                "system_settings": system_settings,
            }
            return render(request, "add_hotel.html", context)

    else:

        forms = villa_Form(instance=instance, user=request.user)
        from masters.models import SystemSettings

        system_settings = (
            SystemSettings.get_settings() if request.user.is_superuser else None
        )

        # Check if villa has bookings
        from customer.models import VillaBooking

        has_bookings = (
            VillaBooking.objects.filter(villa=instance).exists() if instance else False
        )

        context = {
            "form": forms,
            "existing_images": instance.images.all() if instance else None,
            "system_settings": system_settings,
            "has_bookings": has_bookings,
        }

        return render(request, "add_hotel.html", context)


from django.shortcuts import get_object_or_404
from django.db import transaction


@login_required(login_url="login_admin")
def delete_hotel(request, villa_id):

    villa_instance = get_object_or_404(villa, id=villa_id)

    try:
        with transaction.atomic():
            villa_instance.user.delete()  # Delete related user
            villa_instance.delete()  # Delete villa
    except Exception as e:
        # Optionally: log or show error
        print("Error deleting hotel:", e)

    return HttpResponseRedirect(reverse("list_villa"))


from django.db.models import Prefetch


@login_required(login_url="login_admin")
def list_hotel(request):

    data = villa.objects.prefetch_related(
        Prefetch(
            "rooms",
            queryset=villa_rooms.objects.select_related("room_type").prefetch_related(
                "villa_amenities"
            ),
        )
    ).order_by("-id")

    filterset = VillaFilter(request.GET, queryset=data, request=request)
    filtered_bookings = filterset.qs

    # Paginate
    paginator = Paginator(filtered_bookings, 30)  # Show 10 hotels per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "data": page_obj,  # use paginated data
        "filterset": filterset,
        "page_obj": page_obj,
    }

    return render(request, "list_hotel.html", context)


@login_required(login_url="login_admin")
def delete_hotel_image(request, image_id):
    image = get_object_or_404(VillaImage, id=image_id)

    villa_id = image.villa.id  # To redirect back to edit page
    image.delete()

    print("-----------------------------------------")

    print(villa_id)

    return redirect("update_hotel", hotel_id=villa_id)


@login_required(login_url="login_admin")
def add_hotel_rooms(request):
    # For vendors: only allow room addition for Resort and Couple Stay properties
    if request.user.is_service_provider and not request.user.is_superuser:
        # Get user's villas that are Resort or Couple Stay
        user_villas = villa.objects.filter(
            user=request.user,
            is_active=True,
            property_type__name__in=["Resort", "Couple Stay"],
        )

        if not user_villas.exists():
            messages.warning(
                request,
                "Room management is only available for Resort and Couple Stay properties. "
                "Please create a Resort or Couple Stay property first.",
            )
            return redirect("view_villa")

    if request.method == "POST":

        form = villa_rooms_Form(request.POST, request.FILES, user=request.user)

        if form.is_valid():
            instance = form.save(commit=False)

            if request.user.is_superuser:
                # Admin: hotel selected in form by dropdown (already present in form.cleaned_data)
                pass  # already handled by form
            else:
                # Vendor: validate that the selected villa is Resort or Couple Stay
                selected_villa = form.cleaned_data.get("villa")
                if selected_villa:
                    if (
                        selected_villa.property_type
                        and selected_villa.property_type.name
                        not in ["Resort", "Couple Stay"]
                    ):
                        messages.error(
                            request,
                            "Rooms can only be added to Resort or Couple Stay properties. "
                            f"{selected_villa.name} is a {selected_villa.property_type.name}.",
                        )
                        context = {"form": form}
                        return render(request, "add_hotel_rooms.html", context)
                else:
                    # If no villa selected, use first Resort/Couple Stay villa
                    user_villa = villa.objects.filter(
                        user=request.user,
                        is_active=True,
                        property_type__name__in=["Resort", "Couple Stay"],
                    ).first()
                    if not user_villa:
                        return HttpResponse(
                            "You are not linked to any Resort or Couple Stay property. Please add one first.",
                            status=403,
                        )
                    instance.villa = user_villa

            instance.save()
            form.save_m2m()

            for img in request.FILES.getlist("image"):
                villa_roomsImage.objects.create(villa_rooms=instance, image=img)

            messages.success(
                request, f"Room added successfully to {instance.villa.name}!"
            )
            return redirect("list_villa_rooms")

        else:
            print(form.errors)
            # Re-apply villa filtering for vendors in case of errors
            if request.user.is_service_provider and not request.user.is_superuser:
                user_villas = villa.objects.filter(
                    user=request.user,
                    is_active=True,
                    property_type__name__in=["Resort", "Couple Stay"],
                )
                form.fields["villa"].queryset = user_villas
            context = {"form": form}
            return render(request, "add_hotel_rooms.html", context)

    else:

        form = villa_rooms_Form(user=request.user)

        # For vendors: filter villa dropdown to only show Resort/Couple Stay properties
        if request.user.is_service_provider and not request.user.is_superuser:
            user_villas = villa.objects.filter(
                user=request.user,
                is_active=True,
                property_type__name__in=["Resort", "Couple Stay"],
            )
            form.fields["villa"].queryset = user_villas
            form.fields["villa"].help_text = (
                f"Only your Resort and Couple Stay properties are shown. You have {user_villas.count()} property(ies) available."
            )
            if user_villas.count() == 1:
                # Auto-select if only one villa
                form.fields["villa"].initial = user_villas.first()
            elif user_villas.count() == 0:
                form.fields["villa"].help_text = (
                    "You don't have any Resort or Couple Stay properties. Please create one first."
                )

        return render(request, "add_hotel_rooms.html", {"form": form})


@login_required(login_url="login_admin")
def update_hotel_rooms(request, villa_rooms_id):
    instance = get_object_or_404(villa_rooms, id=villa_rooms_id)

    # For vendors: only allow editing rooms for Resort and Couple Stay properties
    if request.user.is_service_provider and not request.user.is_superuser:
        # Check if the room belongs to a Resort or Couple Stay property owned by the user
        if not instance.villa or instance.villa.user != request.user:
            messages.error(request, "You don't have permission to edit this room.")
            return redirect("list_villa_rooms")

        if (
            not instance.villa.property_type
            or instance.villa.property_type.name not in ["Resort", "Couple Stay"]
        ):
            messages.error(
                request,
                "Room editing is only available for Resort and Couple Stay properties.",
            )
            return redirect("list_villa_rooms")

    if request.method == "POST":
        form = villa_rooms_Form(
            request.POST, request.FILES, instance=instance, user=request.user
        )

        if form.is_valid():
            room = form.save(commit=False)

            # Ensure the correct hotel is assigned if user is not a superuser
            if not request.user.is_superuser:
                # Vendor: validate villa is Resort or Couple Stay
                selected_villa = form.cleaned_data.get("villa") or instance.villa
                if selected_villa:
                    if (
                        selected_villa.property_type
                        and selected_villa.property_type.name
                        not in ["Resort", "Couple Stay"]
                    ):
                        messages.error(
                            request,
                            "Rooms can only be assigned to Resort or Couple Stay properties.",
                        )
                        # Re-initialize form with user for proper room type filtering
                        form = villa_rooms_Form(
                            request.POST,
                            request.FILES,
                            instance=instance,
                            user=request.user,
                        )
                        if (
                            request.user.is_service_provider
                            and not request.user.is_superuser
                        ):
                            user_villas = villa.objects.filter(
                                user=request.user,
                                is_active=True,
                                property_type__name__in=["Resort", "Couple Stay"],
                            )
                            form.fields["villa"].queryset = user_villas
                        context = {
                            "form": form,
                            "existing_images": (
                                instance.images.all() if instance else None
                            ),
                        }
                        return render(request, "add_hotel_rooms.html", context)
                    room.villa = selected_villa
                else:
                    # Fallback to existing villa
                    if instance.villa and instance.villa.user != request.user:
                        return HttpResponse(
                            "You don't have permission to edit this room.", status=403
                        )

            room.save()
            form.save_m2m()

            for img in request.FILES.getlist("image"):
                villa_roomsImage.objects.create(villa_rooms=room, image=img)

            messages.success(request, "Room updated successfully!")
            return redirect("list_villa_rooms")
        else:
            print(form.errors)
            # Re-apply villa filtering for vendors in case of errors
            if request.user.is_service_provider and not request.user.is_superuser:
                user_villas = villa.objects.filter(
                    user=request.user,
                    is_active=True,
                    property_type__name__in=["Resort", "Couple Stay"],
                )
                form.fields["villa"].queryset = user_villas
    else:
        form = villa_rooms_Form(instance=instance, user=request.user)

        # For vendors: filter villa dropdown to only show Resort/Couple Stay properties
        if request.user.is_service_provider and not request.user.is_superuser:
            user_villas = villa.objects.filter(
                user=request.user,
                is_active=True,
                property_type__name__in=["Resort", "Couple Stay"],
            )
            form.fields["villa"].queryset = user_villas

    context = {
        "form": form,
        "existing_images": instance.images.all() if instance else None,
    }
    return render(request, "add_hotel_rooms.html", context)


@login_required(login_url="login_admin")
def delete_hotel_rooms(request, villa_rooms_id):
    room = get_object_or_404(villa_rooms, id=villa_rooms_id)

    # For vendors: only allow deleting rooms from their own Resort/Couple Stay properties
    if request.user.is_service_provider and not request.user.is_superuser:
        if not room.villa or room.villa.user != request.user:
            messages.error(request, "You don't have permission to delete this room.")
            return redirect("list_villa_rooms")

        if not room.villa.property_type or room.villa.property_type.name not in [
            "Resort",
            "Couple Stay",
        ]:
            messages.error(
                request,
                "Room deletion is only available for Resort and Couple Stay properties.",
            )
            return redirect("list_villa_rooms")

    room.delete()
    messages.success(request, "Room deleted successfully!")
    return HttpResponseRedirect(reverse("list_villa_rooms"))


@login_required(login_url="login_admin")
def view_hotel_rooms(request, hotel_id):
    villa_instance = get_object_or_404(villa, id=hotel_id)

    # For vendors: only allow viewing rooms from their own Resort/Couple Stay properties
    if request.user.is_service_provider and not request.user.is_superuser:
        if villa_instance.user != request.user:
            messages.error(
                request, "You don't have permission to view rooms for this property."
            )
            return redirect("list_villa_rooms")

        if (
            not villa_instance.property_type
            or villa_instance.property_type.name not in ["Resort", "Couple Stay"]
        ):
            messages.error(
                request,
                "Room viewing is only available for Resort and Couple Stay properties.",
            )
            return redirect("list_villa_rooms")

    data = villa_rooms.objects.filter(villa__id=hotel_id)

    context = {"data": data, "hote_name": villa_instance.name}

    return render(request, "list_hotel_rooms.html", context)


@login_required(login_url="login_admin")
def delete_hotel_room_image(request, image_id):
    image = get_object_or_404(villa_roomsImage, id=image_id)

    villa_rooms_id = image.villa_rooms.id  # To redirect back to edit page
    image.delete()

    print("-----------------------------------------")

    print(villa_rooms_id)

    return redirect("update_villa_rooms", villa_rooms_id=villa_rooms_id)


@login_required(login_url="login_admin")
def list_hotel_rooms(request):
    if request.user.is_superuser:
        # Admin: show all rooms from all villas
        all_rooms = (
            villa_rooms.objects.select_related("villa", "room_type")
            .prefetch_related("villa_amenities", "images")
            .all()
        )

        # Group by villa for display
        villas_with_rooms = {}
        for room in all_rooms:
            if room.villa not in villas_with_rooms:
                villas_with_rooms[room.villa] = []
            villas_with_rooms[room.villa].append(room)

        context = {"villas_with_rooms": villas_with_rooms}
        return render(request, "list_hotel_rooms.html", context)
    else:
        # Vendor: show only rooms from their Resort/Couple Stay properties
        user_villas = villa.objects.filter(
            user=request.user,
            is_active=True,
            property_type__name__in=["Resort", "Couple Stay"],
        )

        if not user_villas.exists():
            messages.info(
                request,
                "You don't have any Resort or Couple Stay properties. "
                "Room management is only available for these property types.",
            )
            return redirect("view_villa")

        # Get all rooms from user's Resort/Couple Stay properties
        all_rooms = (
            villa_rooms.objects.filter(villa__in=user_villas)
            .select_related("villa", "room_type")
            .prefetch_related("villa_amenities", "images")
            .order_by("villa", "room_type")
        )

        # Group by villa for display
        villas_with_rooms = {}
        for room in all_rooms:
            if room.villa not in villas_with_rooms:
                villas_with_rooms[room.villa] = []
            villas_with_rooms[room.villa].append(room)

        context = {"villas_with_rooms": villas_with_rooms}
        return render(request, "list_hotel_rooms.html", context)


from customer.models import *

import openpyxl


def export_bookings_to_excel(queryset):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Villa Bookings"

    # Header row
    headers = [
        "Booking ID",
        "Villa",
        "User",
        "Status",
        "Payment Status",
        "Payment Type",
        "Check-In",
        "Check-Out",
        "Guests",
        "Children",
        "Name",
        "Phone",
        "Email",
        "Base Amount",
        "GST Amount",
        "Total Amount",
        "Booked At",
    ]
    sheet.append(headers)

    # Data rows
    for booking in queryset:
        sheet.append(
            [
                booking.booking_id,
                booking.villa.name if booking.villa else "",
                booking.user.first_name if booking.user else "Guest",
                booking.status,
                booking.payment_status,
                booking.payment_type or "",
                booking.check_in.strftime("%Y-%m-%d"),
                booking.check_out.strftime("%Y-%m-%d"),
                booking.guest_count,
                booking.child_count,
                f"{booking.first_name} {booking.last_name}".strip(),
                booking.phone_number,
                booking.email,
                float(booking.base_amount) if booking.base_amount else 0,
                float(booking.gst_amount) if booking.gst_amount else 0,
                float(booking.total_amount) if booking.total_amount else 0,
                (
                    booking.created_at.strftime("%Y-%m-%d %H:%M")
                    if booking.created_at
                    else ""
                ),
            ]
        )

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="hotel_bookings.xlsx"'
    workbook.save(response)
    return response


@login_required(login_url="login_admin")
def list_hotel_bookings(request):
    # queryset = HotelBooking.objects.filter(payment_status = "paid").order_by('-id') if request.user.is_superuser else HotelBooking.objects.filter(hotel__user=request.user, payment_status = "paid").order_by('-id')
    queryset = queryset = (
        VillaBooking.objects.all().order_by("-id")
        if request.user.is_superuser
        else VillaBooking.objects.filter(villa__user=request.user).order_by("-id")
    )

    filterset = VillaBookingFilter(request.GET, queryset=queryset, request=request)
    filtered_bookings = filterset.qs

    # ✅ Check if "export" is requested
    if "export" in request.GET:
        return export_bookings_to_excel(filtered_bookings)

    paginator = Paginator(filtered_bookings, 30)  # Show 10 hotels per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    total_earning = (
        filtered_bookings.aggregate(total=Sum("hotel_earning"))["total"] or 0
    )

    context = {
        "data": page_obj,
        "filterset": filterset,
        "total_earning": total_earning,
    }
    return render(request, "list_hotel_bookings.html", context)


from django.utils import timezone
from datetime import date


@login_required(login_url="login_admin")
def list_hotel_future_bookings(request):

    today = date.today()

    base_queryset = VillaBooking.objects.filter(
        check_in__gt=today, payment_status="paid"
    )

    if request.user.is_superuser:
        queryset = base_queryset
    else:
        queryset = base_queryset.filter(villa__user=request.user)

    filterset = VillaBookingFilter(request.GET, queryset=queryset, request=request)
    filtered_bookings = filterset.qs

    paginator = Paginator(filtered_bookings, 30)  # Show 10 hotels per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    total_earning = (
        filtered_bookings.aggregate(total=Sum("hotel_earning"))["total"] or 0
    )

    context = {
        "data": page_obj,
        "filterset": filterset,
        "total_earning": total_earning,
    }
    return render(request, "list_hotel_bookings.html", context)


from customer.forms import *


# def send_invoice(request, subject, body, booking_id):
#     booking = get_object_or_404(HotelBooking, id=booking_id)

#     # Generate HTML and PDF
#     html_file_path = generate_invoice_html(booking, request)
#     pdf_file_path = os.path.join(tempfile.gettempdir(), f"invoice_{booking.id}.pdf")
#     html_to_pdf_with_chromium(html_file_path, pdf_file_path)

#     # Create email
#     email = EmailMessage(
#         subject=subject,
#         body=body,
#         from_email='rabbitstay1@gmail.com',
#         to=['pratikgosavi654@gmail.com'],
#     )

#     # Attach PDF
#     if os.path.exists(pdf_file_path):
#         with open(pdf_file_path, 'rb') as f:
#             email.attach(f"invoice_{booking.id}.pdf", f.read(), 'application/pdf')

#     email.send()

#     # Optional: Clean up temp files
#     os.remove(html_file_path)
#     os.remove(pdf_file_path)

#     return HttpResponse("Email with PDF sent successfully!")


@login_required(login_url="login_admin")
def update_hotel_bookings(request, booking_id):

    if request.user.is_superuser:

        instance = VillaBooking.objects.get(id=booking_id)

    else:

        instance = VillaBooking.objects.get(villa__user=request.user, id=booking_id)

    if request.method == "POST":

        print("--------------------")

        form = VillaBookingStatusForm(request.POST, instance=instance)

        print("--------------------")

        if form.is_valid():
            updated_booking = form.save()
            print("--------------------")

            if updated_booking.status == "completed":
                print("Booking has been marked as completed.")

                generate_invoice_pdf(request, updated_booking.id)

            return redirect("list_hotel_bookings")

        else:

            print("------12--------------")

            print(form.errors)
            context = {"form": form}
            return render(request, "update_hotel_bookings.html", context)

    else:

        form = VillaBookingStatusForm(instance=instance, user=request.user)

        context = {"form": form}
        return render(request, "update_hotel_bookings.html", context)


@login_required(login_url="login_admin")
def view_hotel_bookings(request, booking_id):

    instance = HotelBooking.objects.get(id=booking_id)

    form = HotelBookingStatusForm(instance=instance, user=request.user)

    context = {"form": form}
    return render(request, "view_hotel_bookings.html", context)


from django.db.models import Sum

from .filters import *
import openpyxl
from openpyxl.utils import get_column_letter


def export_earning_to_excel(bookings):
    # Create workbook & sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Villa Earnings"

    # Header row
    headers = [
        "Booking ID",
        "Villa",
        "Guest Name",
        "Phone",
        "Email",
        "Check In",
        "Check Out",
        "Nights",
        "Guests",
        "Children",
        "Base Amount",
        "GST",
        "TCS",
        "TDS",
        "Commission",
        "Commission GST",
        "Total Amount",
        "Villa Earning",
        "Status",
        "Created At",
    ]
    ws.append(headers)

    # Data rows
    for booking in bookings:
        nights = (booking.check_out - booking.check_in).days or 1
        ws.append(
            [
                booking.booking_id,
                booking.villa.name if booking.villa else "",
                f"{booking.first_name} {booking.last_name}".strip(),
                booking.phone_number,
                booking.email,
                booking.check_in.strftime("%d-%m-%Y"),
                booking.check_out.strftime("%d-%m-%Y"),
                nights,
                booking.guest_count,
                booking.child_count,
                float(booking.base_amount) if booking.base_amount else 0,
                float(booking.gst_amount) if booking.gst_amount else 0,
                float(booking.tcs_amount) if booking.tcs_amount else 0,
                float(booking.tds_amount) if booking.tds_amount else 0,
                float(booking.commission_amount) if booking.commission_amount else 0,
                float(booking.commission_gst) if booking.commission_gst else 0,
                float(booking.total_amount) if booking.total_amount else 0,
                float(booking.hotel_earning) if booking.hotel_earning else 0,
                booking.get_status_display(),
                (
                    booking.created_at.strftime("%d-%m-%Y %H:%M")
                    if booking.created_at
                    else ""
                ),
            ]
        )

    # Adjust column width
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Prepare response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="villa_earnings.xlsx"'
    wb.save(response)
    return response


@login_required(login_url="login_admin")
def list_hotel_earning(request):

    queryset = (
        VillaBooking.objects.filter(payment_status="paid").order_by("-id")
        if request.user.is_superuser
        else VillaBooking.objects.filter(
            villa__user=request.user, payment_status="paid"
        ).order_by("-id")
    )

    filterset = VillaBookingFilter(request.GET, queryset=queryset, request=request)
    filtered_bookings = filterset.qs
    if "export" in request.GET:
        return export_earning_to_excel(filtered_bookings)
    total_earning = (
        filtered_bookings.aggregate(total=Sum("hotel_earning"))["total"] or 0
    )

    context = {
        "data": filtered_bookings,
        "filterset": filterset,
        "total_earning": total_earning,
    }
    return render(request, "list_hotel_earning.html", context)


# from django.shortcuts import get_object_or_404
# from django.template.loader import render_to_string
# from django.http import HttpResponse
# from xhtml2pdf import pisa
# import io

import io
from django.http import HttpResponse

from django.shortcuts import get_object_or_404
from django.http import HttpResponse

from django.template.loader import render_to_string
import io

import tempfile
from django.template.loader import render_to_string
from django.core.mail import EmailMessage
from django.http import JsonResponse  # optional, if returning a JSON response
from django.conf import settings  #

# from playwright.sync_api import sync_playwright

import os

import base64


# def generate_invoice_html(booking, request):

#     with open(os.path.join(settings.BASE_DIR, 'static/images/Villa_Guru.png'), 'rb') as img_file:
#         logo_base64 = base64.b64encode(img_file.read()).decode('utf-8')

#     html_content = render_to_string('from_owner_to_hotel_invoice.html', {
#         'booking': booking,
#         'request': request,
#         'logo_base64': logo_base64
#     })

#     tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.html', dir=settings.BASE_DIR)
#     tmp.write(html_content.encode('utf-8'))
#     tmp.close()
#     return tmp.name  # return the HTML file path

# from playwright.sync_api import sync_playwright

# def html_to_pdf_with_chromium(html_file_path, output_pdf_path):
#     with sync_playwright() as p:
#         browser = p.chromium.launch(headless=True)
#         page = browser.new_page()
#         page.goto(f'file://{html_file_path}', wait_until='networkidle')
#         page.pdf(path=output_pdf_path, format="A4", print_background=True)
#         browser.close()

# def generate_invoice_pdf(request, booking_id):

#     booking = get_object_or_404(HotelBooking, id=booking_id)

#     html_file_path = generate_invoice_html(booking, request)
#     pdf_file_path = os.path.join(tempfile.gettempdir(), f"invoice_{booking.id}.pdf")

#     html_to_pdf_with_chromium(html_file_path, pdf_file_path)

#     with open(pdf_file_path, 'rb') as f:
#         pdf_data = f.read()

#     # Cleanup
#     os.remove(html_file_path)
#     os.remove(pdf_file_path)

#     return HttpResponse(pdf_data, content_type='application/pdf')


import requests


def generate_invoice_pdf(request, booking_id):
    booking = get_object_or_404(VillaBooking, id=booking_id)
    print("------------------------")
    print(booking.villa.user.email)
    # Generate base64 logo for the template
    with open(
        os.path.join(settings.BASE_DIR, "static/images/Villa_Guru.png"), "rb"
    ) as img_file:
        logo_base64 = base64.b64encode(img_file.read()).decode("utf-8")

    # Render HTML
    html_content = render_to_string(
        "from_owner_to_hotel_invoice.html",
        {"booking": booking, "request": request, "logo_base64": logo_base64},
    )
    subject = "Invoice against Bookinid " + str(booking_id)

    # Generate PDF using html2pdf.app
    response = requests.post(
        "https://api.html2pdf.app/v1/generate",
        json={
            "html": html_content,
            "apiKey": settings.HTML2PDF_API_KEY,
            "options": {"printBackground": True, "margin": "1cm", "pageSize": "A4"},
        },
    )

    if response.status_code != 200:
        return HttpResponse(f"PDF generation failed: {response.text}", status=400)

    pdf_bytes = response.content
    print("------------------------")
    print(booking.villa.user.email)
    # Compose email
    email = EmailMessage(
        subject=subject,
        body="Hi, attached pdf for invoice",
        from_email="Rabbitstay221@gmail.com",
        # to=[booking.villa.user.email],
        to=[booking.villa.user.email],
    )
    email.attach(f"invoice_{booking.id}.pdf", pdf_bytes, "application/pdf")
    email.send()

    return HttpResponse("Email with PDF sent successfully!")


from django.http import JsonResponse
from django.db.models import Q
from datetime import datetime, timedelta
from django.contrib import messages


from collections import defaultdict
import json
from collections import defaultdict


@login_required
def update_hotel_availability(request):
    """
    Manage availability for all properties (Villa, Resort, Couple Stay).
    Shows all properties with booking status and filters.
    Only vendors can access this - not admins.
    """
    from customer.models import VillaBooking

    # Restrict to vendors only
    if request.user.is_superuser:
        messages.info(request, "Availability management is only available for vendors.")
        return redirect("list_villa")

    # Get all properties for vendor
    all_villas = villa.objects.filter(user=request.user, is_active=True).select_related(
        "property_type", "city"
    )

    # Get filter parameters
    filter_status = request.GET.get("status", "all")  # 'all', 'booked', 'available'
    property_type_filter = request.GET.get(
        "property_type", "all"
    )  # 'all', '1', '2', '3'
    villa_id = request.GET.get("villa_id")

    # Filter by property type if specified
    if property_type_filter != "all":
        try:
            all_villas = all_villas.filter(property_type_id=int(property_type_filter))
        except (ValueError, TypeError):
            pass

    # Get selected villa for detailed view
    villa_obj = None
    if villa_id:
        try:
            villa_obj = all_villas.get(id=villa_id)
        except villa.DoesNotExist:
            villa_obj = None

    if request.method == "POST":
        if not villa_obj:
            messages.error(request, "Please select a property first.")
            return HttpResponseRedirect(reverse("update_villa_availability"))

        # Check if this is a bulk management request
        action_type = request.POST.get("action_type")
        if action_type == "bulk_manage":
            return handle_bulk_availability_management(
                request, villa_obj, filter_status, property_type_filter
            )

        selected_date = request.POST.get("selected_date")
        selected_date_obj = datetime.strptime(selected_date, "%Y-%m-%d").date()

        # Handle room-based availability (Resort/Couple Stay)
        property_type_name = (
            villa_obj.property_type.name if villa_obj.property_type else None
        )
        if property_type_name in ["Resort", "Couple Stay"]:
            for room in villa_rooms.objects.filter(villa=villa_obj):
                field_name = f"availability_{room.id}"
                count = request.POST.get(field_name)

                if count is not None and count != "":
                    count_int = int(count)
                    # Save/update room availability
                    # If count is 0, room is marked as booked (offline)
                    RoomAvailability.objects.update_or_create(
                        room=room,
                        date=selected_date_obj,
                        defaults={"available_count": count_int},
                    )
            messages.success(request, f"Availability updated for {selected_date}")
        else:
            # Handle villa availability (whole villa)
            is_open = request.POST.get("is_open", "true") == "true"

            # Check if there's an active booking for this date
            # If booking exists, don't allow vendor to open it (booking takes priority)
            has_active_booking = VillaBooking.objects.filter(
                villa=villa_obj,
                check_in__lte=selected_date_obj,
                check_out__gt=selected_date_obj,
                status__in=["confirmed", "checked_in", "pending"],
            ).exists()

            if has_active_booking and is_open:
                messages.warning(
                    request,
                    f"Cannot mark as available - there is an active booking on {selected_date}. "
                    "Cancel the booking first to make it available.",
                )
            else:
                VillaAvailability.objects.update_or_create(
                    villa=villa_obj,
                    date=selected_date_obj,
                    defaults={"is_open": is_open},
                )
                if is_open:
                    messages.success(
                        request, f"Villa marked as available for {selected_date}"
                    )
                else:
                    messages.success(
                        request, f"Villa marked as unavailable for {selected_date}"
                    )

        return HttpResponseRedirect(
            reverse("update_villa_availability")
            + f"?villa_id={villa_obj.id}&status={filter_status}&property_type={property_type_filter}"
        )

    # Get all bookings for all properties to determine booking status
    from customer.models import VillaBooking
    from datetime import date as date_type, timedelta
    from django.db.models import Q, Count, Exists, OuterRef

    # Get current date for checking availability
    today = date_type.today()

    # Annotate each villa with booking status
    villas_with_status = []
    for v in all_villas:
        # Check if villa has any active bookings
        active_bookings = VillaBooking.objects.filter(
            villa=v,
            status__in=["confirmed", "checked_in", "pending"],
            check_in__lte=today + timedelta(days=365),  # Check next year
            check_out__gte=today,
        )

        has_active_booking = active_bookings.exists()

        # Get next booking if any
        next_booking = active_bookings.order_by("check_in").first()

        villas_with_status.append(
            {
                "villa": v,
                "has_booking": has_active_booking,
                "next_booking": next_booking,
                "booking_count": active_bookings.count(),
            }
        )

    # Apply status filter
    if filter_status == "booked":
        villas_with_status = [v for v in villas_with_status if v["has_booking"]]
    elif filter_status == "available":
        villas_with_status = [v for v in villas_with_status if not v["has_booking"]]

    # Prepare context for selected villa (if any)
    availability_data = {}
    events = []
    booked_dates = []
    rooms = []

    if villa_obj:
        # Get property type
        property_type_name = (
            villa_obj.property_type.name if villa_obj.property_type else None
        )

        if property_type_name in ["Resort", "Couple Stay"]:
            # Room-based availability
            rooms = villa_rooms.objects.filter(villa=villa_obj).select_related(
                "room_type"
            )
            raw_availability = RoomAvailability.objects.filter(room__villa=villa_obj)

            # Build JSON: { "2025-07-05": { "8": 5, "9": 3 } }
            availability_data = defaultdict(dict)
            for entry in raw_availability:
                availability_data[entry.date.isoformat()][
                    str(entry.room.id)
                ] = entry.available_count

            # Combine room data by date for single event display
            grouped = defaultdict(list)
            booked_rooms_grouped = defaultdict(list)
            for entry in raw_availability:
                if entry.available_count == 0:
                    # Marked as booked (offline)
                    booked_rooms_grouped[entry.date.isoformat()].append(
                        f"{entry.room.room_type.name if entry.room.room_type else 'Room'}: Booked (Offline)"
                    )
                else:
                    grouped[entry.date.isoformat()].append(
                        f"{entry.room.room_type.name if entry.room.room_type else 'Room'}: {entry.available_count} rooms"
                    )

            # Create availability events (blue color for available)
            availability_events = [
                {
                    "title": "<br>".join(labels),
                    "start": date,
                    "color": "#007bff",
                    "display": "block",
                }
                for date, labels in grouped.items()
            ]
            events.extend(availability_events)

            # Create booked (offline) events (yellow/orange color)
            offline_booked_events = [
                {
                    "title": "<br>".join(labels),
                    "start": date,
                    "color": "#ffc107",
                    "display": "block",
                }
                for date, labels in booked_rooms_grouped.items()
            ]
            events.extend(offline_booked_events)
        else:
            # Villa-based availability
            raw_availability = VillaAvailability.objects.filter(villa=villa_obj)
            availability_data = defaultdict(dict)

            # First, get all bookings to know which dates are booked
            bookings = VillaBooking.objects.filter(
                villa=villa_obj, status__in=["confirmed", "checked_in", "pending"]
            )

            # Create a set of all booked dates from bookings
            booked_dates_from_bookings = set()
            for booking in bookings:
                current_date = booking.check_in
                while current_date < booking.check_out:
                    booked_dates_from_bookings.add(current_date.isoformat())
                    current_date += timedelta(days=1)

            # Process availability records - add ALL dates to availability_data
            # This ensures we have data for all dates, even if no record exists
            for entry in raw_availability:
                availability_data[entry.date.isoformat()] = {"is_open": entry.is_open}
                date_str = entry.date.isoformat()

                # Check if this date has a booking
                has_booking = date_str in booked_dates_from_bookings

                if not entry.is_open:
                    if has_booking:
                        # Booking exists - will be shown as red/orange background from booking events below
                        # Don't add yellow event, let booking event show
                        pass
                    else:
                        # Manually closed (vendor closed it, no booking)
                        events.append(
                            {
                                "title": f"Unavailable (Manually Set)",
                                "start": date_str,
                                "color": "#ffc107",  # Yellow for manually closed
                                "display": "background",  # Show as background
                                "allDay": True,
                            }
                        )

        # Get bookings for selected villa (if not already fetched above)
        if villa_obj.property_type and villa_obj.property_type.name != "Villa":
            bookings = VillaBooking.objects.filter(
                villa=villa_obj, status__in=["confirmed", "checked_in", "pending"]
            )
        elif not "bookings" in locals():
            bookings = VillaBooking.objects.filter(
                villa=villa_obj, status__in=["confirmed", "checked_in", "pending"]
            )

        # Create a set of all booked dates
        booked_dates_set = set()

        for booking in bookings:
            # Add all dates between check_in and check_out (exclusive of check_out)
            current_date = booking.check_in
            while current_date < booking.check_out:
                booked_dates_set.add(current_date.isoformat())
                current_date += timedelta(days=1)

            # Add booking event for the calendar
            payment_type_label = (
                "Online" if booking.payment_type == "online" else "Offline"
            )
            # Use different color for offline bookings to distinguish them
            event_color = (
                "#dc3545" if booking.payment_type == "online" else "#ff9800"
            )  # Orange for offline

            events.append(
                {
                    "title": f"Booked ({payment_type_label}): {booking.booking_id or 'N/A'} - {booking.guest_count} guest(s)",
                    "start": booking.check_in.isoformat(),
                    "end": booking.check_out.isoformat(),
                    "color": event_color,  # Red for online, Orange for offline
                    "display": "background",  # Show as background color
                    "allDay": True,
                }
            )

        booked_dates = list(booked_dates_set)

    # Get room booking status for Resort/Couple Stay properties
    rooms_with_status = []
    if villa_obj:
        property_type_name = (
            villa_obj.property_type.name if villa_obj.property_type else None
        )
        if property_type_name in ["Resort", "Couple Stay"]:
            from customer.models import BookingRoom

            # Get date range from query params (default to today + 30 days)
            check_date_from = request.GET.get("check_date_from")
            check_date_to = request.GET.get("check_date_to")

            if check_date_from and check_date_to:
                try:
                    check_from = datetime.strptime(check_date_from, "%Y-%m-%d").date()
                    check_to = datetime.strptime(check_date_to, "%Y-%m-%d").date()
                except (ValueError, TypeError):
                    check_from = date_type.today()
                    check_to = date_type.today() + timedelta(days=30)
            else:
                check_from = date_type.today()
                check_to = date_type.today() + timedelta(days=30)

            # Get all bookings for these rooms in the date range
            room_bookings = BookingRoom.objects.filter(
                booking__villa=villa_obj,
                booking__status__in=["confirmed", "checked_in", "pending"],
                booking__check_in__lt=check_to,
                booking__check_out__gt=check_from,
            ).select_related("booking", "room")

            # Calculate booked dates per room
            room_booked_dates = defaultdict(set)
            for room_booking in room_bookings:
                current_date = max(check_from, room_booking.booking.check_in)
                end_date = min(check_to, room_booking.booking.check_out)
                while current_date < end_date:
                    room_booked_dates[room_booking.room_id].add(current_date)
                    current_date += timedelta(days=1)

            # Check each room's status
            for room in rooms:
                # Check offline bookings (available_count = 0)
                offline_booked_dates = RoomAvailability.objects.filter(
                    room=room,
                    date__gte=check_from,
                    date__lt=check_to,
                    available_count=0,
                ).values_list("date", flat=True)

                # Check online bookings
                online_booked_dates = room_booked_dates.get(room.id, set())

                # Combine all booked dates
                all_booked_dates = set(offline_booked_dates) | online_booked_dates

                # Calculate total days in range
                total_days = (check_to - check_from).days
                booked_days_count = len(all_booked_dates)

                # Determine status
                if booked_days_count == total_days:
                    status = "fully_booked"
                    status_label = "Fully Booked"
                elif booked_days_count > 0:
                    status = "partially_booked"
                    status_label = (
                        f"Partially Booked ({booked_days_count}/{total_days} days)"
                    )
                else:
                    status = "available"
                    status_label = "Available"

                rooms_with_status.append(
                    {
                        "room": room,
                        "status": status,
                        "status_label": status_label,
                        "booked_days": booked_days_count,
                        "total_days": total_days,
                        "booked_dates": sorted(all_booked_dates),
                    }
                )

    # Get property types for filter dropdown
    from masters.models import property_type

    property_types = property_type.objects.all()

    context = {
        "all_villas": villas_with_status,
        "villa_obj": villa_obj,
        "rooms": rooms,
        "rooms_with_status": rooms_with_status,
        "availability_json": json.dumps(availability_data),
        "events": json.dumps(events),
        "booked_dates_json": json.dumps(booked_dates),
        "filter_status": filter_status,
        "property_type_filter": property_type_filter,
        "property_types": property_types,
        "selected_villa_id": villa_obj.id if villa_obj else None,
        "check_date_from": request.GET.get("check_date_from", ""),
        "check_date_to": request.GET.get("check_date_to", ""),
    }

    return render(request, "update_hotel_availability.html", context)


def handle_bulk_availability_management(
    request, villa_obj, filter_status, property_type_filter
):
    """
    Handle bulk availability management with status options:
    - available: Mark as available
    - offline_booked: Mark as offline booked (close availability)
    - unavailable: Mark as unavailable (manually closed)
    """
    from datetime import timedelta
    from customer.models import VillaBooking

    bulk_from_date = request.POST.get("bulk_from_date")
    bulk_to_date = request.POST.get("bulk_to_date")
    bulk_status = request.POST.get(
        "bulk_status"
    )  # available, offline_booked, unavailable

    try:
        from_date_obj = datetime.strptime(bulk_from_date, "%Y-%m-%d").date()
        to_date_obj = datetime.strptime(bulk_to_date, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        messages.error(request, "Invalid date range")
        return HttpResponseRedirect(
            reverse("update_villa_availability")
            + f"?villa_id={villa_obj.id}&status={filter_status}&property_type={property_type_filter}"
        )

    if from_date_obj > to_date_obj:
        messages.error(request, "'From' date must be before 'To' date.")
        return HttpResponseRedirect(
            reverse("update_villa_availability")
            + f"?villa_id={villa_obj.id}&status={filter_status}&property_type={property_type_filter}"
        )

    property_type_name = (
        villa_obj.property_type.name if villa_obj.property_type else None
    )

    updated_count = 0
    skipped_count = 0
    current_date = from_date_obj

    if property_type_name in ["Resort", "Couple Stay"]:
        # Room-based availability
        bulk_rooms = request.POST.getlist("bulk_rooms[]")
        rooms_to_update = (
            villa_rooms.objects.filter(villa=villa_obj, id__in=bulk_rooms)
            if bulk_rooms
            else villa_rooms.objects.filter(villa=villa_obj)
        )

        while current_date <= to_date_obj:
            for room in rooms_to_update:
                # Check if room has active bookings
                has_booking = (
                    VillaBooking.objects.filter(
                        villa=villa_obj,
                        booking_type="selected_rooms",
                        status__in=["confirmed", "checked_in", "pending"],
                        check_in__lte=current_date,
                        check_out__gt=current_date,
                    )
                    .filter(booked_rooms__room=room)
                    .exists()
                )

                if has_booking and bulk_status == "available":
                    skipped_count += 1
                    continue

                # Set availability based on status
                if bulk_status == "available":
                    # Set to 1 room available (or get from room default)
                    available_count = 1
                elif bulk_status == "offline_booked":
                    # Mark as booked offline (0 available)
                    available_count = 0
                else:  # unavailable
                    # Mark as unavailable (0 available)
                    available_count = 0

                RoomAvailability.objects.update_or_create(
                    room=room,
                    date=current_date,
                    defaults={"available_count": available_count},
                )
                updated_count += 1

            current_date += timedelta(days=1)

        status_label = {
            "available": "available",
            "offline_booked": "offline booked",
            "unavailable": "unavailable",
        }.get(bulk_status, "updated")

        messages.success(
            request,
            f"Updated {updated_count} room availability records to '{status_label}'. "
            f"{skipped_count} skipped (active bookings).",
        )

    else:
        # Villa-based availability
        while current_date <= to_date_obj:
            # Check if villa has active bookings
            has_booking = VillaBooking.objects.filter(
                villa=villa_obj,
                booking_type="whole_villa",
                status__in=["confirmed", "checked_in", "pending"],
                check_in__lte=current_date,
                check_out__gt=current_date,
            ).exists()

            if has_booking and bulk_status == "available":
                skipped_count += 1
                current_date += timedelta(days=1)
                continue

            # Set availability based on status
            if bulk_status == "available":
                is_open = True
            else:  # offline_booked or unavailable
                is_open = False

            VillaAvailability.objects.update_or_create(
                villa=villa_obj,
                date=current_date,
                defaults={"is_open": is_open},
            )
            updated_count += 1
            current_date += timedelta(days=1)

        status_label = {
            "available": "available",
            "offline_booked": "offline booked",
            "unavailable": "unavailable",
        }.get(bulk_status, "updated")

        messages.success(
            request,
            f"Updated {updated_count} dates to '{status_label}'. "
            f"{skipped_count} skipped (active bookings).",
        )

    return HttpResponseRedirect(
        reverse("update_villa_availability")
        + f"?villa_id={villa_obj.id}&status={filter_status}&property_type={property_type_filter}"
    )


from datetime import datetime, timedelta
from collections import defaultdict
import json


@login_required
def update_from_to_hotel_availability(request):
    """
    Bulk update availability for date range.
    Supports both Villa (whole property) and Resort/Couple Stay (room-based).
    Only vendors can access this - not admins.
    """
    # Restrict to vendors only
    if request.user.is_superuser:
        messages.info(request, "Availability management is only available for vendors.")
        return redirect("list_villa")

    # Get villa from GET or POST
    villa_id = request.GET.get("villa_id") or request.POST.get("villa_id")
    villas_list = villa.objects.filter(user=request.user, is_active=True)

    if villa_id:
        try:
            villa_obj = villas_list.get(id=villa_id)
        except villa.DoesNotExist:
            messages.error(request, "Property not found.")
            return HttpResponseRedirect(reverse("update_villa_availability"))
    else:
        villa_obj = villas_list.first() if villas_list.exists() else None

    if not villa_obj:
        messages.error(request, "You are not linked to any property.")
        return redirect("vendor_dashboard")

    if request.method == "POST":
        from_date = request.POST.get("from_date")
        to_date = request.POST.get("to_date")

        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            messages.error(request, "Invalid date range")
            return HttpResponseRedirect(
                reverse("update_villa_availability") + f"?villa_id={villa_obj.id}"
            )

        if from_date_obj > to_date_obj:
            messages.error(request, "'From' date must be before 'To' date.")
            return HttpResponseRedirect(
                reverse("update_villa_availability") + f"?villa_id={villa_obj.id}"
            )

        property_type_name = (
            villa_obj.property_type.name if villa_obj.property_type else None
        )
        updated_count = 0
        current_date = from_date_obj

        if property_type_name in ["Resort", "Couple Stay"]:
            # Room-based availability
            from customer.models import BookingRoom

            while current_date <= to_date_obj:
                # Check if rooms have bookings on this date
                for room in villa_rooms.objects.filter(villa=villa_obj):
                    field_name = f"availability_{room.id}"
                    count = request.POST.get(field_name)

                    if count is not None and count != "":
                        # Check existing bookings for this room
                        has_booking = BookingRoom.objects.filter(
                            booking__villa=villa_obj,
                            room=room,
                            booking__status__in=["confirmed", "checked_in", "pending"],
                            booking__check_in__lte=current_date,
                            booking__check_out__gt=current_date,
                        ).exists()

                        if not has_booking:
                            RoomAvailability.objects.update_or_create(
                                room=room,
                                date=current_date,
                                defaults={"available_count": int(count)},
                            )
                            updated_count += 1
                current_date += timedelta(days=1)
        else:
            # Villa-based availability
            while current_date <= to_date_obj:
                # Check if villa has bookings on this date
                has_booking = VillaBooking.objects.filter(
                    villa=villa_obj,
                    booking_type="whole_villa",
                    status__in=["confirmed", "checked_in", "pending"],
                    check_in__lte=current_date,
                    check_out__gt=current_date,
                ).exists()

                is_open = (
                    request.POST.get("is_open", "true") == "true" and not has_booking
                )

                VillaAvailability.objects.update_or_create(
                    villa=villa_obj,
                    date=current_date,
                    defaults={"is_open": is_open},
                )
                updated_count += 1
                current_date += timedelta(days=1)

        messages.success(
            request,
            f"Availability updated for {updated_count} days from {from_date} to {to_date}.",
        )

        return HttpResponseRedirect(
            reverse("update_villa_availability") + f"?villa_id={villa_obj.id}"
        )

    return HttpResponseRedirect(reverse("update_villa_availability"))


@login_required(login_url="login_admin")
def manage_villa_pricing(request):
    """
    Main view for managing villa pricing with calendar interface.
    For Villa properties: manages whole villa pricing
    For Resort/Couple Stay: manages individual room pricing
    """
    # Get villa_id and room_id from GET parameters
    villa_id = request.GET.get("villa_id")
    room_id = request.GET.get("room_id")

    if request.user.is_superuser:
        # Admin can view any villa
        if villa_id:
            try:
                villa_obj = villa.objects.get(id=villa_id)
            except villa.DoesNotExist:
                messages.error(request, "Villa not found.")
                return redirect("list_villa")
        else:
            # If no villa_id, get first villa or show all villas
            villas_list = villa.objects.all().order_by("name")
            if villas_list.exists():
                villa_obj = villas_list.first()
            else:
                messages.error(request, "No villas found.")
                return redirect("list_villa")
    else:
        # Vendor can see their own villas - get first one or use selected villa
        villas_list = villa.objects.filter(user=request.user, is_active=True)
        villa_id = request.GET.get("villa_id")

        if villa_id:
            try:
                villa_obj = villas_list.get(id=villa_id)
            except villa.DoesNotExist:
                messages.error(request, "Selected villa not found.")
                villa_obj = villas_list.first() if villas_list.exists() else None
        else:
            villa_obj = villas_list.first() if villas_list.exists() else None

        if not villa_obj:
            messages.error(
                request, "You are not linked to any villa. Please add a villa first."
            )
            return redirect("add_villa")

    # Determine if this is a room-based property (Resort/Couple Stay)
    is_room_based = villa_obj.property_type and villa_obj.property_type.name in [
        "Resort",
        "Couple Stay",
    ]
    selected_room = None

    if is_room_based:
        # Get rooms for this villa
        rooms_list = (
            villa_rooms.objects.filter(villa=villa_obj)
            .select_related("room_type")
            .order_by("room_type__name")
        )

        # Get selected room or first room
        if room_id:
            try:
                selected_room = rooms_list.get(id=room_id)
            except villa_rooms.DoesNotExist:
                selected_room = rooms_list.first() if rooms_list.exists() else None
        else:
            selected_room = rooms_list.first() if rooms_list.exists() else None

        if not selected_room and rooms_list.exists():
            # If no room selected but rooms exist, redirect to first room
            return redirect(
                f"{request.path}?villa_id={villa_obj.id}&room_id={rooms_list.first().id}"
            )
    else:
        rooms_list = []

    if request.method == "POST":
        # Handle single date pricing update
        selected_date = request.POST.get("selected_date")
        price = request.POST.get("price_per_night")
        room_id_post = request.POST.get("room_id")  # For room-based pricing

        if selected_date and price:
            try:
                selected_date_obj = datetime.strptime(selected_date, "%Y-%m-%d").date()
                price_decimal = Decimal(price)

                # Check if villa/room has bookings on this date
                from customer.models import VillaBooking, BookingRoom

                if is_room_based and room_id_post:
                    # Room-based booking check
                    room_obj = villa_rooms.objects.get(id=room_id_post)
                    has_booking = BookingRoom.objects.filter(
                        booking__villa=villa_obj,
                        room=room_obj,
                        booking__status__in=["confirmed", "checked_in"],
                        booking__check_in__lte=selected_date_obj,
                        booking__check_out__gt=selected_date_obj,
                    ).exists()
                else:
                    # Villa-based booking check
                    has_booking = VillaBooking.objects.filter(
                        villa=villa_obj,
                        booking_type="whole_villa",
                        status__in=["confirmed", "checked_in"],
                        check_in__lte=selected_date_obj,
                        check_out__gt=selected_date_obj,
                    ).exists()

                if has_booking:
                    messages.error(
                        request,
                        f"⚠️ Cannot change price for {selected_date} because there is an existing booking for this date. "
                        "Please contact admin if you need to update the price.",
                    )
                else:
                    if is_room_based and room_id_post:
                        # Room-based pricing
                        room_obj = villa_rooms.objects.get(id=room_id_post)
                        RoomPricing.objects.update_or_create(
                            room=room_obj,
                            date=selected_date_obj,
                            defaults={"price_per_night": price_decimal},
                        )
                        messages.success(
                            request, f"Room price updated for {selected_date}"
                        )
                    else:
                        # Villa-based pricing
                        VillaPricing.objects.update_or_create(
                            villa=villa_obj,
                            date=selected_date_obj,
                            defaults={"price_per_night": price_decimal},
                        )
                        messages.success(request, f"Price updated for {selected_date}")
            except (ValueError, TypeError) as e:
                messages.error(request, f"Invalid date or price: {str(e)}")

        # Redirect with villa_id and room_id to maintain selection
        from django.http import HttpResponseRedirect
        from django.urls import reverse

        redirect_url = reverse("manage_villa_pricing")
        params = []
        if villa_obj:
            params.append(f"villa_id={villa_obj.id}")
        if is_room_based and room_id_post:
            params.append(f"room_id={room_id_post}")

        if params:
            redirect_url += "?" + "&".join(params)

        return HttpResponseRedirect(redirect_url)

    # Get existing pricing data
    if is_room_based and selected_room:
        # Room-based pricing
        raw_pricing = RoomPricing.objects.filter(room=selected_room)
        default_price = float(selected_room.price_per_night or 0)
        weekend_price = None
        if villa_obj.weekend_percentage and selected_room.price_per_night:
            weekend_multiplier = Decimal(1) + (villa_obj.weekend_percentage / 100)
            weekend_price = round(selected_room.price_per_night * weekend_multiplier, 2)
    else:
        # Villa-based pricing
        raw_pricing = VillaPricing.objects.filter(villa=villa_obj)
        default_price = float(villa_obj.price_per_night or 0)
        weekend_price = None
        if villa_obj.weekend_percentage and villa_obj.price_per_night:
            weekend_multiplier = Decimal(1) + (villa_obj.weekend_percentage / 100)
            weekend_price = round(villa_obj.price_per_night * weekend_multiplier, 2)

    # Build JSON: { "2025-07-05": 1500.00 }
    pricing_data = {}
    for entry in raw_pricing:
        pricing_data[entry.date.isoformat()] = float(entry.price_per_night)

    # Get all villas for dropdown (admin sees all, vendor sees their own)
    villas_list = []
    if request.user.is_superuser:
        villas_list = list(villa.objects.filter(is_active=True).order_by("name"))
    else:
        villas_list = list(
            villa.objects.filter(user=request.user, is_active=True).order_by("name")
        )

    context = {
        "villa": villa_obj,
        "villas_list": villas_list,
        "rooms_list": rooms_list if is_room_based else [],
        "selected_room": selected_room,
        "is_room_based": is_room_based,
        "default_price": default_price,
        "weekend_percentage": float(villa_obj.weekend_percentage or 0),
        "weekend_price": float(weekend_price) if weekend_price else None,
        "pricing_json": json.dumps(pricing_data),
        "pricing_list": raw_pricing.order_by("-date")[:50],  # Show last 50 entries
    }

    return render(request, "manage_villa_pricing.html", context)


@login_required(login_url="login_admin")
def bulk_update_villa_pricing(request):
    """
    Bulk update pricing for a date range.
    Only vendors can access this - not admins.
    """
    # Restrict to vendors only
    if request.user.is_superuser:
        messages.info(request, "Pricing management is only available for vendors.")
        return redirect("list_villa")

    # Get first villa for vendor (or selected one if multiple)
    villas_list = villa.objects.filter(user=request.user, is_active=True)
    villa_id = request.GET.get("villa_id")

    if villa_id:
        try:
            villa_obj = villas_list.get(id=villa_id)
        except villa.DoesNotExist:
            villa_obj = villas_list.first() if villas_list.exists() else None
    else:
        villa_obj = villas_list.first() if villas_list.exists() else None

    if not villa_obj:
        messages.error(
            request, "You are not linked to any villa. Please add a villa first."
        )
        return redirect("add_villa")

    # Determine if this is a room-based property
    is_room_based = villa_obj.property_type and villa_obj.property_type.name in [
        "Resort",
        "Couple Stay",
    ]
    room_id = request.GET.get("room_id") or request.POST.get("room_id")
    selected_room = None

    if is_room_based and room_id:
        try:
            selected_room = villa_rooms.objects.get(id=room_id, villa=villa_obj)
        except villa_rooms.DoesNotExist:
            selected_room = None

    if request.method == "POST":
        from_date = request.POST.get("from_date")
        to_date = request.POST.get("to_date")
        price = request.POST.get("price_per_night")
        # Get villa_id and room_id from POST if provided, otherwise use GET
        villa_id = request.POST.get("villa_id") or request.GET.get("villa_id")
        room_id = request.POST.get("room_id") or request.GET.get("room_id")

        # Re-fetch villa_obj if villa_id changed
        if villa_id:
            try:
                villa_obj = villas_list.get(id=villa_id)
                is_room_based = (
                    villa_obj.property_type
                    and villa_obj.property_type.name in ["Resort", "Couple Stay"]
                )
            except villa.DoesNotExist:
                villa_obj = villas_list.first() if villas_list.exists() else None

        # Re-fetch room if room_id provided
        if is_room_based and room_id:
            try:
                selected_room = villa_rooms.objects.get(id=room_id, villa=villa_obj)
            except villa_rooms.DoesNotExist:
                selected_room = None

        if not all([from_date, to_date, price]):
            messages.error(request, "All fields are required.")
            redirect_url = reverse("manage_villa_pricing")
            params = []
            if villa_obj:
                params.append(f"villa_id={villa_obj.id}")
            if is_room_based and room_id:
                params.append(f"room_id={room_id}")
            if params:
                redirect_url += "?" + "&".join(params)
            return HttpResponseRedirect(redirect_url)

        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            price_decimal = Decimal(price)
        except (ValueError, TypeError) as e:
            messages.error(request, f"Invalid date or price: {str(e)}")
            redirect_url = reverse("manage_villa_pricing")
            params = []
            if villa_obj:
                params.append(f"villa_id={villa_obj.id}")
            if is_room_based and room_id:
                params.append(f"room_id={room_id}")
            if params:
                redirect_url += "?" + "&".join(params)
            return HttpResponseRedirect(redirect_url)

        if from_date_obj > to_date_obj:
            messages.error(request, "'From' date must be before 'To' date.")
            redirect_url = reverse("manage_villa_pricing")
            params = []
            if villa_obj:
                params.append(f"villa_id={villa_obj.id}")
            if is_room_based and room_id:
                params.append(f"room_id={room_id}")
            if params:
                redirect_url += "?" + "&".join(params)
            return HttpResponseRedirect(redirect_url)

        # Update pricing for each date in range
        current_date = from_date_obj
        updated_count = 0
        skipped_count = 0
        from customer.models import VillaBooking

        while current_date <= to_date_obj:
            # Check if villa/room has bookings on this date
            if is_room_based and selected_room:
                # Room-based booking check
                from customer.models import BookingRoom

                has_booking = BookingRoom.objects.filter(
                    booking__villa=villa_obj,
                    room=selected_room,
                    booking__status__in=["confirmed", "checked_in"],
                    booking__check_in__lte=current_date,
                    booking__check_out__gt=current_date,
                ).exists()
            else:
                # Villa-based booking check
                has_booking = VillaBooking.objects.filter(
                    villa=villa_obj,
                    booking_type="whole_villa",
                    status__in=["confirmed", "checked_in"],
                    check_in__lte=current_date,
                    check_out__gt=current_date,
                ).exists()

            if not has_booking:
                if is_room_based and selected_room:
                    RoomPricing.objects.update_or_create(
                        room=selected_room,
                        date=current_date,
                        defaults={"price_per_night": price_decimal},
                    )
                else:
                    VillaPricing.objects.update_or_create(
                        villa=villa_obj,
                        date=current_date,
                        defaults={"price_per_night": price_decimal},
                    )
                updated_count += 1
            else:
                skipped_count += 1
            current_date += timedelta(days=1)

        if skipped_count > 0:
            messages.warning(
                request,
                f"Pricing updated for {updated_count} days from {from_date} to {to_date}. "
                f"{skipped_count} date(s) were skipped because they have existing bookings.",
            )
        else:
            messages.success(
                request,
                f"Pricing successfully updated for {updated_count} days from {from_date} to {to_date}.",
            )

    # Redirect with villa_id to maintain selection
    redirect_url = reverse("manage_villa_pricing")
    params = []
    if villa_obj:
        params.append(f"villa_id={villa_obj.id}")
    # Check if this is room-based pricing
    room_id = request.GET.get("room_id") or request.POST.get("room_id")
    if room_id:
        params.append(f"room_id={room_id}")
    if params:
        redirect_url += "?" + "&".join(params)
    return HttpResponseRedirect(redirect_url)


@login_required(login_url="login_admin")
def delete_villa_pricing(request, pricing_id):
    """
    Delete a specific date pricing entry.
    Only vendors can access this - not admins.
    Prevents deletion if there are bookings for that date.
    """
    # Restrict to vendors only
    if request.user.is_superuser:
        messages.info(request, "Pricing management is only available for vendors.")
        return redirect("list_villa")

    try:
        pricing_obj = VillaPricing.objects.get(id=pricing_id, villa__user=request.user)
        villa_obj = pricing_obj.villa

        # Check if villa has bookings on this date
        from customer.models import VillaBooking

        has_booking = VillaBooking.objects.filter(
            villa=pricing_obj.villa,
            status__in=["confirmed", "checked_in"],
            check_in__lte=pricing_obj.date,
            check_out__gt=pricing_obj.date,
        ).exists()

        if has_booking:
            messages.error(
                request,
                f"⚠️ Cannot delete pricing for {pricing_obj.date} because there is an existing booking for this date. "
                "Please contact admin if you need to delete the pricing.",
            )
        else:
            pricing_obj.delete()
            messages.success(
                request, f"Pricing for {pricing_obj.date} has been deleted."
            )
    except VillaPricing.DoesNotExist:
        messages.error(request, "Pricing entry not found.")
        villa_obj = None

    # Redirect with villa_id to maintain selection
    redirect_url = reverse("manage_villa_pricing")
    params = []
    if villa_obj:
        params.append(f"villa_id={villa_obj.id}")
    # Check if this is room-based pricing
    room_id = request.GET.get("room_id") or request.POST.get("room_id")
    if room_id:
        params.append(f"room_id={room_id}")
    if params:
        redirect_url += "?" + "&".join(params)
    return HttpResponseRedirect(redirect_url)
